# -*- coding: utf-8 -*-
"""
Created on Tue Nov  3 14:57:21 2020

@author: palav
"""


import pandas as pd
import xlsxwriter

pd.set_option("display.max_colwidth", 200)
from openpyxl import Workbook,load_workbook

#-----DATASET-ANSVERS
wb = load_workbook("C:/Users/palav/OneDrive/Masaüstü/bitirme_tezi/dataset.xlsx")
ws = wb.active
#------SCORES
scores="C:/Users/palav/OneDrive/Masaüstü/bitirme_tezi/ansvers.xlsx"
workbook = xlsxwriter.Workbook(scores)
worksheet = workbook.add_worksheet()
worksheetl=load_workbook(scores)
sheet=worksheetl.active

total_columns=int(ws.max_column)
total_rows=int(ws.max_row)
liste=[]
ansver_true=[]

array_ansvers_excel_row=[]
array_ansvers_excel_row_new=[]
array_ansvers_excel_col_new=[]
#total_columns_ansver=int(sheet.max_column)
#total_rows_ansver=int(sheet.max_row)

def ansver_excel():
        #-----------------CEVAPLAR İÇİN EXCEL DOSYASI OLUŞTURMA------------------#  

    #-------BAŞLIK
    rows=ws.iter_rows(min_row=1, min_col=2, max_row=1, max_col=total_columns)
    for row in rows:
        for cell in row:
            array_ansvers_excel_row.append(cell)
        
        
    #------SÜTUN    
    array_ansvers_excel_col=[]
    rows=ws.iter_rows(min_row=1, min_col=1, max_row=total_rows, max_col=1)
    for row in rows:
        for cell in row:
            array_ansvers_excel_col.append(cell)
        #--------------STUDENTS--------------------#
    row_name=0
    col_name=1
    
    for i in range(len(array_ansvers_excel_row)):
         array_ansvers_excel_row_new.append(array_ansvers_excel_row[i].value)
    for i in array_ansvers_excel_row_new:
         worksheet.write(row_name, col_name, i)
         col_name+=1
     #-------------------ID---------------------#
    row_name=0
    col_name=0
     
    for i in range(len(array_ansvers_excel_col)):
         array_ansvers_excel_col_new.append(array_ansvers_excel_col[i].value)    
    for i in array_ansvers_excel_col_new:        
         worksheet.write(row_name, col_name,i)
         row_name+=1

    worksheetl.save(scores)
    #-----------------------!EXCEL OLUŞTURULDU---------------------#  

#TABLOYU OLUŞTURDUM
ansver_excel()

def clean_shred(arraycleanshred):
    liste.clear()#bizi bitirdin beeeeeeeeee
    array_cleanshred=[]
    array_cleanshred=arraycleanshred 
    ansvers_array_one=[] #orijinal metin    
    ansvers_array_onetoone=[] #temizlenmiş metin    
    results=[]
       
    len_array_cleanshred=(len(array_cleanshred))
    #len_results=len_ansvers_array       
    #-------------METİN TEMİZLEME VE LİSTEYE ATMA-----------------------#
    for i in range(len_array_cleanshred):
        ansvers_array_one.append(array_cleanshred[i].value)#SAĞLAM        
        #len_ansvers_array_one=len(ansvers_array_one)
        news_df = pd.DataFrame({'document':ansvers_array_one})      
        # removing everything except alphabets` 
        news_df['clean_doc'] = news_df['document'].str.replace("[^abcçdefgğhıijklmnoöpqrsştuüvwxy-z.ABCÇDEFGĞHIİJKLMNOÖPQRSŞTUÜVWXY-Z.#]", " ") 
        # removing short words 
        news_df['clean_doc'] = news_df['clean_doc'].apply(lambda x: ' '.join([w for w in x.split() if len(w)>3])) 
        # make all text lowercase 
        news_df['clean_doc'] = news_df['clean_doc'].apply(lambda x: x.lower())
        new_dataset=news_df['clean_doc'] 
    #--------------------------------------------------------------------#
    #-----------------DATAFRAME TO ARRAY---------------------------------#  
    
    for j in new_dataset :#SAĞLAM
        ansvers_array_onetoone.append(j)
        len_ansvers_array_onetoone=len(ansvers_array_onetoone)
    #--------------------------------------------------------------------#
    #---HER BÖLÜNME İŞLEMİNİ İÇEREN BİR ALT LİSTELER KÜMESİ OLUŞTURUR----#           
    for k in range(0,len_ansvers_array_onetoone):                          
        results.append(ansvers_array_onetoone[k])
        for l in results:            
            section=results[k].split('. ')
        liste.append(section)
    
    return liste    

       
def search(keyword,prekeyword,suffix_keyword):       
    ansver_true=[]
    key=keyword
    key1=prekeyword
    key2=suffix_keyword
    for listindex in liste:        
        for listindex_index in listindex:
            if key in listindex_index:                                                
                if key1 in listindex_index:
                    if key2 in listindex_index:
                        #----listede bulduğuz cevap tekrar ederse listeye eklenmesini önlüyor
                        if liste.index(listindex) in ansver_true:
                            print()
                        else:
                            ansver_true.append(liste.index(listindex))                           
                   
    return ansver_true
    


def question_1(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    ansvers_array=[] #dosyadan istediğimiz satırdaki metin
    
#-------------------1.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
#ha buraya bir else eklenecek öğrenci cevabı yoksa dizi index değeri none olarak kaydedilecek
    rows=ws.iter_rows(min_row=2, min_col=2, max_row=2, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#    
    clean_shred(ansvers_array)       

    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)
    
    #--1.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için

    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için
    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=1
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
       
   
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!1.SORU İŞLEMLERİ TAMAMDIR-------------------------------------# 


#--------------------------2. SORU------------------------------#  
def question_2(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    ansvers_array=[] #dosyadan istediğimiz satırdaki metin
    
#-------------------2.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=3, min_col=2, max_row=3, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--2.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=2
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
    
   
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1         
    
    worksheetl.save(scores) 

#------------------------!2.SORU İŞLEMLERİ TAMAMDIR-------------------
 
#--------------------------3. SORU------------------------------#  
def question_3(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    ansvers_array=[] #dosyadan istediğimiz satırdaki metin
    
#-------------------3.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=4, min_col=2, max_row=4, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--3.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=3
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
   
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
     
#------------------------!3.SORU İŞLEMLERİ TAMAMDIR-------------------

#--------------------------4. SORU------------------------------#  
def question_4(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    ansvers_array=[] #dosyadan istediğimiz satırdaki metin
    
#-------------------4.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=5, min_col=2, max_row=5, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #----------4.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=4
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
    
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!4.SORU İŞLEMLERİ TAMAMDIR-------------------
#--------------------------5. SORU------------------------------#  
def question_5(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    ansvers_array=[] #dosyadan istediğimiz satırdaki metin

#-------------------5.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=6, min_col=2, max_row=6, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--5.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=5
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
    
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!5.SORU İŞLEMLERİ TAMAMDIR-------------------
#--------------------------6. SORU------------------------------#  


def question_6(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    
    ansvers_array=[]
#-------------------6.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=7, min_col=2, max_row=7, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--6.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=6
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
   
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!6.SORU İŞLEMLERİ TAMAMDIR-------------------

#---------------------7. SORU İŞLEMLERİ-----------------------
def question_7(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    
    ansvers_array=[]
#-------------------7.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=8, min_col=2, max_row=8, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--7.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=7
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
    
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!7.SORU İŞLEMLERİ TAMAMDIR-------------------

#--------------------8.SORU İŞLEMLERİ----------------------

def question_8(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    
    ansvers_array=[]
#-------------------8.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=9, min_col=2, max_row=9, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--8.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=8
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
   
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    
#------------------------!8.SORU İŞLEMLERİ TAMAMDIR-------------------
#--------------------9.SORU İŞLEMLERİ----------------------

def question_9(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    
    ansvers_array=[]
#-------------------9.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=10, min_col=2, max_row=10, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--8.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=9
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
    
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    #workbook.close()#son save'den sonra kullanılacak
#------------------------!9.SORU İŞLEMLERİ TAMAMDIR-------------------

#--------------------10.SORU İŞLEMLERİ----------------------

def question_10(keyword1,prekeyword1,suffix_keyword1,
                keyword2,prekeyword2,suffix_keyword2,
                keyword3,prekeyword3,suffix_keyword3,
                keyword4,prekeyword4,suffix_keyword4,
                keyword5,prekeyword5,suffix_keyword5):
    
    
    ansvers_array=[]
#-------------------10.SORU İÇİN TÜM ÖĞRENCİLERİN CEVAPLARI---------------#
    rows=ws.iter_rows(min_row=11, min_col=2, max_row=11, max_col=total_columns)
    for row in rows:
        for cell in row:
            ansvers_array.append(cell)   
    
#------------------------------------------------------------------------#
    #-------------metini temizledim ve parçaladım------------------------#
    
    clean_shred(ansvers_array)       
    
    #---------------HER CEVAP İÇİN ANAHTAR KELİMELERİ ARA 1.keyword----------------#    
    question1=search(keyword1,prekeyword1,suffix_keyword1)
    question2=search(keyword2,prekeyword2,suffix_keyword2)
    question3=search(keyword3,prekeyword3,suffix_keyword3)                 
    question4=search(keyword4,prekeyword4,suffix_keyword4)
    question5=search(keyword5,prekeyword5,suffix_keyword5)

    #--10.SORU CEVAPLARINI OLUŞTURDUĞUMUZ ANSVERS DOSYASINA ATMAK--#
    #----------------------------------#
    #---------------------ANAHTAR KELİMELER GÖRE ÖĞRENCİYİ NOTLAMA--------------#
    
    array1_1=[]    
    #ilk anahtar kelime için koşulu sağlayan öğrencilerin puanını 1 yazıyor----#
    for j in range(total_columns-1):
            if j in question1:
                array1_1.append(1)
                j+=1
            else:
                array1_1.append(0)
                j+=1
    
    #ikinci anahtar kelime için

    array1_2=[]
    for j in range(total_columns-1):
            if j in question2:
                array1_2.append(1)
                j+=1
            else:
                array1_2.append(0)
                j+=1
    
        
    #ücüncü anahtar kelime için
    
    array1_3=[]
    for j in range(total_columns-1):
            if j in question3:
                array1_3.append(1)
                j+=1
            else:
                array1_3.append(0)
                j+=1

    #dördüncü anahtar kelime için     
    
    array1_4=[]
    for j in range(total_columns-1):
            if j in question4:
                array1_4.append(1)
                j+=1
            else:
                array1_4.append(0)
                j+=1

    #besinci anahtar kelime için    
    array1_5=[]
    for j in range(total_columns-1):
            if j in question5:
                array1_5.append(1)
                j+=1
            else:
                array1_5.append(0)
                j+=1
    
    #------------------------------------------------------------------------------------#                      
    #-------------------ÖĞRENCİNİN TOPLAM NOTUNU EXCELLE YAZDIRMA------------------------#  
    row_name=10
    col_name=1         
    array_ansvers_=[]
    for i in range(total_columns-1):
        x=array1_1[i]+array1_2[i]+array1_3[i]+array1_4[i]+array1_5[i]
        array_ansvers_.append(x)
        
    
    for i in array_ansvers_:
        worksheet.write(row_name, col_name,i)
        col_name+=1 
        
    
    worksheetl.save(scores)    
    workbook.close()#son save'den sonra kullanılacak
#------------------------!10.SORU İŞLEMLERİ TAMAMDIR-------------------


question_1("bilgisayar","hesaplama","tarih",
            "bilgisayar","mekanik","dişli",
            "cebir","boole","ikili sayı sistemi",
            "algoritmalar","program","yürütme",
            "programlama dilleri","ada","makine dili")

question_2("ikili sayı sistemi","transistör","elektronik",
            "ikili","cebit","mantık",
            "doğru","mantık","yanlış",
            "entegre","sayısal","kapı",
            "keşif","bilgisayar","teorem")

question_3("analiz","problem","karar",
            "sonlu","adım","kesin",
            "şekil","sorun","şema",
            "standart","basit","sembol",
            "gerçek kod","kaba kod","programlama dili")

question_4("programlama","ilk","dili",
            "seviye","düşük","makine",
            "seviye","yüksek","dil",
            "yorumlayıcı","derleyici","yaklaşım",
            "ymnelimli","nesne","dil")

question_5("işletim","ilk","sistem",
            "ıbm","unix","geliştirme",
            "ms-dos","windows","microsoft",
            "kullanıcı","tür","gömülü",
            "çekirdek","donanım","kabuk")

question_6("kaynak","haberleşme","paylaş",
            "topoloji","ağ","mantık",
            "ıpv6","ıpv4","geçiş",
            "tcp/ıp","protokol","paket",
            "sosyal ağlar","büyük veri","bulut")

question_7("mühendislik","yazılım","bilgisayar",
            "modelleme","geliştirme","tasarım",
            "çözümleme","sistem","uml",
            "sınıf","senaryo","durum",
            "agile","metodoloji","scrum")

question_8("algoritmalar","veri yapıları","hesaplama",
            "basit","veri yapısı","tip",
            "ileri","veri modeli","tür",
            "bellek","çalışma hızı","zaman",
            "algoritma","problem","saklanma")

question_9("ilişki","veri","disk",
            "gizlilik","güvenlik","erişim",
            "ilişkisel","yönetim sistemi","yaklaşım",
            "sorgu","sql","veritabanı",
            "nosql","büyük veri","ilişkisiz")

question_10("donanım","programlama","komut",
            "sayı","ikili","sistemi",
            "onaltılı","sekizli","sistem",
            "klavye","bilgisayar","kodlama",
            "kodlama","karakter","tablo")